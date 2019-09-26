import sys
import boto3
import csv
from openpyxl import Workbook
client = boto3.client('route53')

r53=client.list_hosted_zones_by_name()
sys.stdout=open("output_hostedzones.csv","w")
print ("The number of HostedZones is",len (r53["HostedZones"]))
try:
    print ("%s,%s,%s,%s" %("Name","Id","DNSName","HostedZoneId"))
    for r in r53['HostedZones']:
            rname = r['Name']
            rid = r['Id']
            #dnsname = r['DNSName']
            #zoneid = r['HostedZoneId']
            print("%s,%s"%(rname,rid))#,dnsname,zoneid))
except KeyError as k:
    print ('I got a KeyError - reason "%s" ' % str(k))
except Exception as e:
                print(e,'this is exception')
sys.stdout.close()
################################# Listing the domains##############################
'''
client = boto3.client('route53domains')
domain_list = client.list_domains()
sys.stdout = open ("output_domains.csv","w")
print "The number of domains is",len (domain_list["Domains"])
print "DomainName\tAutoRenew\tTransferLock\tExpiry"
try:
  for domain in domain_list['Domains']:
     print "{dname}\t{arenew}\t{tlock}\t{expiry}".format(
                dname = domain['DomainName'],
                arenew = domain['AutoRenew'],
                tlock = domain['DNSName'],
                expiry = domain['Expiry']
           )
except KeyError as k:
    print 'I got a KeyError - reason "%s"' % str(k)
except Exception as e:
                print(e,'this is exception')
sys.stdout.close()
'''
################################# Converting csv to excel #########################
wb = Workbook()
sheet1 = wb.create_sheet('HostedZones',0)
#sheet2 = wb.create_sheet('Domains',1)
sheet1 = wb['HostedZones']
with open('./output_hostedzones.csv', 'r') as f:
    for row in csv.reader(f,delimiter = "\t"):
        sheet1.append(row)
'''
sheet2 = wb['Domains']
with open('./output_domains.csv', 'r') as f:
    for row in csv.reader(f,delimiter = "\t"):
        sheet2.append(row)
'''
